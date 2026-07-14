import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TREATMENT_LABELS = {
    "mitigate": "Reducir",
    "transfer": "Transferir",
    "accept": "Aceptar",
    "avoid": "Evitar",
}

RISK_LABELS = {
    "critical": "Crítico",
    "high": "Alto",
    "medium": "Medio",
    "low": "Bajo",
}

PRIORITY_LABELS = {
    "immediate": "Inmediato",
    "short_term": "Corto plazo",
    "medium_term": "Mediano plazo",
    "long_term": "Largo plazo",
}

COLS = [
    ("ID de riesgo", 15),
    ("Nivel de riesgo", 20),
    ("ModalidadTratamiento", 25),
    ("Acción a realizar", 60),
    ("Recursos necesarios", 40),
    ("Responsable", 30),
    ("Plazo / Fecha limite", 20),
    ("Prioridad", 20),
    ("Reducción esperada", 25),
    ("Riesgo residual esperado", 25),
]

def _header_row(ws, row: int, cols: list[str]):
    fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    font = Font(color="F9FAFB", bold=True, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        top=Side(style="thin", color="374151"),
        bottom=Side(style="thin", color="374151"),
        left=Side(style="thin", color="374151"),
        right=Side(style="thin", color="374151"),
    )
    for col_idx, title in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border

def _data_row(ws, row: int, data: list, alt: bool = False):
    fill = PatternFill(start_color="F9FAFB" if alt else "FFFFFF", end_color="F9FAFB" if alt else "FFFFFF", fill_type="solid")
    font = Font(color="111827", size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
    )
    for col_idx, val in enumerate(data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = fill
        cell.font = font
        cell.border = border
        if col_idx in [4, 5]:
            cell.alignment = align_left
        else:
            cell.alignment = align_center

def build_treatments_excel(treatments: list[dict], project_name: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan de Tratamiento"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    title_cell = ws.cell(row=1, column=1, value="PLAN DE TRATAMIENTO")
    title_cell.font = Font(size=16, bold=True, color="1F2937")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
    meta_cell = ws.cell(row=2, column=1, value=f"Proyecto: {project_name}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    meta_cell.font = Font(size=11, color="4B5563", italic=True)
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 10

    _header_row(ws, 4, [c[0] for c in COLS])
    ws.row_dimensions[4].height = 42

    row = 5
    for i, t in enumerate(treatments):
        desc = []
        if t.get("title"): desc.append(t["title"])
        if t.get("description"): desc.append(t["description"])
        accion = "\n".join(desc) if desc else "-"

        due_date = t.get("due_date")
        date_str = due_date.strftime("%d/%m/%Y") if due_date else "-"

        reduction = t.get("expected_risk_reduction")
        reduction_str = f"{reduction}%" if reduction is not None else "-"

        t_type = t.get("treatment_type", "")
        t_type_str = TREATMENT_LABELS.get(t_type, t_type.capitalize()) if t_type else "-"

        priority_str = PRIORITY_LABELS.get(t.get("priority"), str(t.get("priority") or "-"))

        r_level = t.get("risk_level", "")
        r_level_str = RISK_LABELS.get(r_level, r_level.capitalize()) if r_level else "-"

        residual = t.get("residual_level", "")
        residual_str = RISK_LABELS.get(residual, residual.capitalize()) if residual else "-"

        effort_map = {"low": "Esfuerzo bajo", "medium": "Esfuerzo medio", "high": "Esfuerzo alto"}
        recursos = effort_map.get(t.get("effort"), "-")

        _data_row(ws, row, [
            t.get("risk_code") or "-",
            r_level_str,
            t_type_str,
            accion,
            recursos,
            t.get("owner_name") or "-",
            date_str,
            priority_str,
            reduction_str,
            residual_str
        ], alt=(i % 2 == 1))

        lines = max(len(accion.split("\n")), 1)
        ws.row_dimensions[row].height = max(30, lines * 16)
        row += 1

    for j, (_, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
