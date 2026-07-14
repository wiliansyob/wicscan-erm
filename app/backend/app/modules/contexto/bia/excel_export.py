"""
TFM Excel export — 4 sheets:
  Tabla 11  BIA-1  Clasificacion de procesos por criticidad operativa
  Tabla 12  BIA-2  Estimacion de impacto y umbrales de tolerancia
  Tabla 15  Matriz de identificacion de activos y vulnerabilidades
  Tabla 16  Matriz de vulnerabilidades contextual (escenarios + riesgo + negocio)
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─── Palette ──────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold: bool = False, size: int = 10, color: str = "1A1A2E") -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


HEADER_FILL   = _fill("2C3E7A")
SUBHEAD_FILL  = _fill("4A6FA5")
ALT_FILL      = _fill("EEF2F9")
WHITE_FILL    = _fill("FFFFFF")
TITLE_FILL    = _fill("1A2E5A")

# Criticality (process)
CRIT_COLORS = {
    "critical":  ("C0392B", "FFFFFF"),
    "important": ("E67E22", "FFFFFF"),
    "support":   ("27AE60", "FFFFFF"),
}
CRIT_LABELS = {
    "critical":  "Critico",
    "important": "Importante",
    "support":   "Soporte",
}

# Severity (findings)
SEV_COLORS = {
    "critical": ("7B0000", "FFFFFF"),
    "high":     ("C0392B", "FFFFFF"),
    "medium":   ("E67E22", "FFFFFF"),
    "low":      ("27AE60", "FFFFFF"),
    "info":     ("7F8C8D", "FFFFFF"),
}
SEV_LABELS = {
    "critical": "Critico",
    "high":     "Alto",
    "medium":   "Medio",
    "low":      "Bajo",
    "info":     "Info",
}

# Risk level
RISK_COLORS = {
    "critical": ("7B0000", "FFFFFF"),
    "high":     ("C0392B", "FFFFFF"),
    "medium":   ("E67E22", "FFFFFF"),
    "low":      ("27AE60", "FFFFFF"),
}
RISK_LABELS = {
    "critical": "Critico",
    "high":     "Alto",
    "medium":   "Medio",
    "low":      "Bajo",
}

# Finding type
FINDING_TYPE_LABELS = {
    "vulnerability":     "Vulnerabilidad",
    "code_smell":        "Code smell",
    "bug":               "Bug",
    "security_hotspot":  "Hotspot",
}

REVENUE_LABEL = {
    ">50":   ">50%",
    "20-50": "20-50%",
    "<20":   "<20%",
}
MANUAL_LABELS = {
    "documented": "Si (documentada)",
    "partial":    "Parcial",
    "none":       "No disponible",
}

STATUS_LABELS = {
    "open":            "Abierto",
    "confirmed":       "Confirmado",
    "false_positive":  "Falso positivo",
    "resolved":        "Resuelto",
    "accepted":        "Aceptado",
    "wont_fix":        "No se corregira",
    "pending":         "Pendiente",
}

IMPACT_LEVEL_LABELS = {
    # Spanish (stored in DB)
    "muy alto":  "Muy Alto",
    "alto":      "Alto",
    "medio":     "Medio",
    "bajo":      "Bajo",
    "muy bajo":  "Muy Bajo",
    # English (legacy)
    "very_high": "Muy Alto",
    "high":      "Alto",
    "medium":    "Medio",
    "low":       "Bajo",
    "very_low":  "Muy Bajo",
}
IMPACT_LEVEL_COLORS = {
    # Spanish (stored in DB)
    "muy alto":  ("7B0000", "FFFFFF"),
    "alto":      ("C0392B", "FFFFFF"),
    "medio":     ("E67E22", "FFFFFF"),
    "bajo":      ("27AE60", "FFFFFF"),
    "muy bajo":  ("2ECC71", "FFFFFF"),
    # English (legacy)
    "very_high": ("7B0000", "FFFFFF"),
    "high":      ("C0392B", "FFFFFF"),
    "medium":    ("E67E22", "FFFFFF"),
    "low":       ("27AE60", "FFFFFF"),
    "very_low":  ("2ECC71", "FFFFFF"),
}

PROB_LEVEL_LABELS = {
    # Spanish (stored in DB)
    "muy alta":  "Muy Alta",
    "alta":      "Alta",
    "media":     "Media",
    "baja":      "Baja",
    "muy baja":  "Muy Baja",
    # English (legacy)
    "very_high": "Muy Alta",
    "high":      "Alta",
    "medium":    "Media",
    "low":       "Baja",
    "very_low":  "Muy Baja",
}

TREATMENT_TYPE_LABELS = {
    "mitigate":  "Mitigar",
    "accept":    "Aceptar",
    "transfer":  "Transferir",
    "avoid":     "Evitar",
}


# ─── Row helpers ──────────────────────────────────────────────────────────────

def _section_title(ws: Any, row: int, text: str, ncols: int, fill: PatternFill = TITLE_FILL) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill
    c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c.alignment = _center()
    c.border = _border()


def _note_row(ws: Any, row: int, text: str, ncols: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill("FFF9C4")
    c.font = Font(name="Calibri", italic=True, size=9, color="555555")
    c.alignment = _left()
    c.border = _border()


def _header_row(ws: Any, row: int, values: list[str]) -> None:
    for j, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=j, value=val)
        c.fill = HEADER_FILL
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.alignment = _center()
        c.border = _border()


def _data_row(ws: Any, row: int, values: list[Any], alt: bool = False, left_cols: set[int] | None = None) -> None:
    fill = ALT_FILL if alt else WHITE_FILL
    left_cols = left_cols or {1}
    for j, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=j, value=val)
        c.fill = fill
        c.font = _font()
        c.alignment = _left() if j in left_cols else _center()
        c.border = _border()


def _colored_cell(ws: Any, row: int, col: int, value: str, bg: str, fg: str) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg)
    c.font = Font(name="Calibri", bold=True, size=10, color=fg)
    c.alignment = _center()
    c.border = _border()


# ─── BIA-1 ────────────────────────────────────────────────────────────────────

def _build_bia1(ws: Any, processes: list[dict], project_name: str) -> None:
    ws.title = "BIA-1 Procesos"
    COLS = [
        ("Proceso de negocio",            32),
        ("Responsable",                   20),
        ("Sistemas / Activos de soporte", 42),
        ("% Ingresos dependientes",       20),
        ("Alternativa manual",            20),
        ("Compromisos contractuales",     22),
        ("Prioridad / Criticidad",        20),
    ]
    N = len(COLS)
    row = 1

    _section_title(ws, row, "TFM - WicScan Risk Manager", N)
    ws.row_dimensions[row].height = 22; row += 1

    _section_title(ws, row,
        "Tabla 11. BIA-1 Clasificacion de procesos por criticidad operativa",
        N, fill=SUBHEAD_FILL)
    ws.row_dimensions[row].height = 20; row += 1

    _note_row(ws, row,
        f"Proyecto: {project_name}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        N)
    ws.row_dimensions[row].height = 16; row += 1

    _header_row(ws, row, [c[0] for c in COLS])
    ws.row_dimensions[row].height = 36; row += 1

    for i, p in enumerate(processes):
        crit_key = p.get("criticality") or "support"
        rev      = REVENUE_LABEL.get(p.get("revenue_dependency") or "", "-")
        manual   = MANUAL_LABELS.get(p.get("manual_alternative") or "none", "No disponible")
        cc       = p.get("contractual_commitment")
        contrac  = "Si" if (cc and isinstance(cc, dict) and cc.get("exists")) else "No"

        _data_row(ws, row, [
            p.get("process_name") or p.get("name", ""),
            p.get("owner") or "-",
            p.get("asset_names") or "-",
            rev, manual, contrac,
            CRIT_LABELS.get(crit_key, crit_key.capitalize()),
        ], alt=(i % 2 == 1))

        hex_bg, hex_fg = CRIT_COLORS.get(crit_key, ("888888", "FFFFFF"))
        _colored_cell(ws, row, 7, CRIT_LABELS.get(crit_key, crit_key.capitalize()), hex_bg, hex_fg)
        ws.row_dimensions[row].height = 28; row += 1

    _data_row(ws, row, [""] * N); ws.row_dimensions[row].height = 22; row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
    leg = ws.cell(row=row, column=1,
        value="LEYENDA:  Critico = core sin tolerancia  |  Importante = afecta ingresos  |  Soporte = auxiliar")
    leg.fill = _fill("F5F5F5")
    leg.font = Font(name="Calibri", size=9, italic=True, color="444444")
    leg.alignment = _left()

    for j, (_, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


# ─── BIA-2 ────────────────────────────────────────────────────────────────────

def _build_bia2(ws: Any, processes: list[dict], project_name: str) -> None:
    ws.title = "BIA-2 Impacto"
    COLS = [
        ("Proceso critico",          28),
        ("Clasificacion",            16),
        ("I(RPO) EUR",               14),
        ("I(RTO) EUR",               14),
        ("I(MTPD) EUR",              14),
        ("SN activa (RGPD)",         16),
        ("MTPD (horas)",             13),
        ("RTO (horas)",              12),
        ("RPO (horas)",              12),
        ("Parametros de calculo",    38),
    ]
    N = len(COLS)
    row = 1

    _section_title(ws, row, "TFM - WicScan Risk Manager", N)
    ws.row_dimensions[row].height = 22; row += 1

    _section_title(ws, row,
        "Tabla 12. BIA-2 Estimacion de impacto y umbrales de tolerancia por proceso critico",
        N, fill=SUBHEAD_FILL)
    ws.row_dimensions[row].height = 20; row += 1

    _note_row(ws, row,
        f"Proyecto: {project_name}   |   I(t)=Perdida estimada EUR   |   SN=Sanciones normativas   |   "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", N)
    ws.row_dimensions[row].height = 16; row += 1

    _header_row(ws, row, [c[0] for c in COLS])
    ws.row_dimensions[row].height = 40; row += 1

    eur_fmt = '#,##0.00'
    for i, p in enumerate(processes):
        crit_key = p.get("criticality") or "support"
        bia      = p.get("bia") or {}
        sn       = "SI" if bia.get("sn_active") else "No"

        bd     = bia.get("breakdown") or {}
        params = bd.get("params", {}) if isinstance(bd, dict) else {}

        # Use pre-computed BIA values stored in the database (I(RPO)=impact_2h, I(RTO)=impact_8h, I(MTPD)=impact_24h)
        val_rpo  = float(bia.get("impact_2h")  or 0)
        val_rto  = float(bia.get("impact_8h")  or 0)
        val_mtpd = float(bia.get("impact_24h") or 0)

        parts: list[str] = []
        if params.get("hourly_revenue"):
            parts.append(f"Ingreso/h: {params['hourly_revenue']:.0f} EUR")
        if params.get("num_staff_affected"):
            parts.append(f"Empleados: {params['num_staff_affected']}")
        if params.get("infra_cost_per_hour"):
            parts.append(f"Infra/h: {params['infra_cost_per_hour']:.0f} EUR")
        if params.get("contractual_penalty_per_hour"):
            parts.append(f"Penalizacion/h: {params['contractual_penalty_per_hour']:.0f} EUR")

        _data_row(ws, row, [
            p.get("process_name") or p.get("name", ""),
            CRIT_LABELS.get(crit_key, crit_key.capitalize()),
            float(val_rpo),
            float(val_rto),
            float(val_mtpd),
            sn,
            bia.get("mtpd_hours"), bia.get("rto_hours"), bia.get("rpo_hours"),
            " | ".join(parts) if parts else "Parametros no configurados",
        ], alt=(i % 2 == 1))

        for col_idx in [3, 4, 5]:
            nc = ws.cell(row=row, column=col_idx)
            nc.number_format = eur_fmt
            nc.alignment = _center()

        hex_bg, hex_fg = CRIT_COLORS.get(crit_key, ("888888", "FFFFFF"))
        _colored_cell(ws, row, 2, CRIT_LABELS.get(crit_key, crit_key.capitalize()), hex_bg, hex_fg)

        sn_cell = ws.cell(row=row, column=6)
        if bia.get("sn_active"):
            sn_cell.fill = _fill("FFCDD2")
            sn_cell.font = Font(name="Calibri", bold=True, size=10, color="C62828")
        sn_cell.alignment = _center()

        i_mtpd = float(val_mtpd)
        i24_cell = ws.cell(row=row, column=5)
        if i_mtpd > 50000:
            i24_cell.fill = _fill("FFCDD2")
            i24_cell.font = Font(name="Calibri", bold=True, size=10, color="C62828")
        elif i_mtpd > 10000:
            i24_cell.fill = _fill("FFE0B2")
            i24_cell.font = Font(name="Calibri", bold=True, size=10, color="E65100")

        ws.row_dimensions[row].height = 28; row += 1

    _data_row(ws, row, [""] * N); ws.row_dimensions[row].height = 22; row += 2

    _section_title(ws, row, "GLOSARIO DE COLUMNAS", N, fill=SUBHEAD_FILL)
    ws.row_dimensions[row].height = 18; row += 1
    for term, desc in [
        ("I(RPO)",    "Perdida acumulada al alcanzar el Objetivo de Punto de Recuperacion"),
        ("I(RTO)",    "Perdida acumulada al alcanzar el Objetivo de Tiempo de Recuperacion"),
        ("I(MTPD)",   "Perdida acumulada al alcanzar el Periodo Maximo Tolerable de Interrupcion"),
        ("SN activa", "SI: incidente implica notificacion a AEPD en 72h (RGPD Art.33)"),
        ("MTPD",      "Periodo maximo de interrupcion tolerable antes de impacto irreversible"),
        ("RTO",       "Objetivo de tiempo de recuperacion"),
        ("RPO",       "Objetivo de punto de recuperacion — maxima antiguedad de los datos"),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        c = ws.cell(row=row, column=1, value=f"  {term}: {desc}")
        c.fill = _fill("F5F5F5"); c.font = Font(name="Calibri", size=9, color="333333")
        c.alignment = _left(); c.border = _border()
        ws.row_dimensions[row].height = 16; row += 1

    for j, (_, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


# ─── Tabla 14 — Matriz de identificacion de activos ─────────

def _build_tabla14(ws: Any, assets: list[dict], project_name: str) -> None:
    ws.title = "T14 Identificacion Activos"
    COLS = [
        ("ID Activo",              15),
        ("Nombre del activo",      28),
        ("Categoria",              16),
        ("Responsable",            22),
        ("Detalle",                40),
        ("Proceso que soporta",    26),
    ]
    N = len(COLS)
    row = 1

    _section_title(ws, row, "TFM - WicScan Risk Manager", N)
    ws.row_dimensions[row].height = 22; row += 1

    _section_title(ws, row,
        "Tabla 14. Matriz de Identificacion de activos",
        N, fill=SUBHEAD_FILL)
    ws.row_dimensions[row].height = 20; row += 1

    _note_row(ws, row,
        f"Proyecto: {project_name}   |   {len(assets)} activos registrados   |   "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", N)
    ws.row_dimensions[row].height = 16; row += 1

    _header_row(ws, row, [c[0] for c in COLS])
    ws.row_dimensions[row].height = 36; row += 1

    for i, a in enumerate(assets):
        # Generar un ID simplificado basado en index si no tenemos asset_code. Usamos ACT-00X
        asset_id_str = f"ACT-{(i+1):03d}"
        
        # Combinar detalle (URL, IP, Descripcion)
        det_parts = []
        if a.get("url"): det_parts.append(a.get("url"))
        if a.get("ip_address"): det_parts.append(a.get("ip_address"))
        if a.get("description"): det_parts.append(a.get("description"))
        detalle = " | ".join(det_parts) if det_parts else "-"

        # Responsable (tecnico o negocio)
        resp = a.get("technical_owner") or a.get("business_owner") or "-"

        _data_row(ws, row, [
            asset_id_str,
            a.get("name") or "-",
            a.get("asset_type") or "-",
            resp,
            detalle,
            a.get("process_names") or "-",
        ], alt=(i % 2 == 1), left_cols={1, 2, 4, 5, 6})

        ws.row_dimensions[row].height = 22; row += 1

    for j, (_, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


# ─── Tabla 16 — Matriz de registros de activos con sus hallazgos ─────────

def _build_tabla16_findings(ws: Any, findings: list[dict], asset_map: dict, project_name: str) -> None:
    """
    1 fila = 1 hallazgo.
    Inventario tecnico completo: activo x vulnerabilidad.
    findings dict keys:
        finding_code, title, finding_type, category, cwe, owasp_category,
        cvss_score, severity, scanner, confidence, file_path, line_start,
        status, first_detected_at,
        asset_name, asset_type, asset_criticality, asset_owner
    """
    ws.title = "T16 Registros Activos-Hallazgos"
    COLS = [
        ("Activos",             24),
        ("Hallazgos",           12),
        ("Titulo hallazgo",     42),
        ("Tipo de hallazgo",    18),
        ("Fuente",              14),
        ("Referencia tecnica",  30),
        ("Severidad",           12),
        ("Confianza",           10),
    ]
    N = len(COLS)
    row = 1

    _section_title(ws, row, "TFM - WicScan Risk Manager", N)
    ws.row_dimensions[row].height = 22; row += 1

    _section_title(ws, row,
        "Tabla 16. Matriz de registros de activos con sus hallazgos",
        N, fill=SUBHEAD_FILL)
    ws.row_dimensions[row].height = 20; row += 1

    _note_row(ws, row,
        f"Proyecto: {project_name}   |   {len(findings)} hallazgos activos   |   "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", N)
    ws.row_dimensions[row].height = 16; row += 1

    _header_row(ws, row, [c[0] for c in COLS])
    ws.row_dimensions[row].height = 36; row += 1

    for i, f in enumerate(findings):
        sev_key   = (f.get("severity") or "info").lower()
        location  = f.get("file_path") or ""
        if location and f.get("line_start"):
            location = f"{location}:{f['line_start']}"
        det_at    = f.get("first_detected_at")
        if det_at and hasattr(det_at, "strftime"):
            det_str = det_at.strftime("%d/%m/%Y")
        elif det_at:
            det_str = str(det_at)[:10]
        else:
            det_str = "-"

        cvss = f.get("cvss_score")
        conf = f.get("confidence")

        _data_row(ws, row, [
            asset_map.get(f.get("asset_name"), "ACT-000"),
            f.get("finding_code") or "-",
            f.get("title") or "-",
            FINDING_TYPE_LABELS.get(f.get("finding_type") or "", f.get("finding_type") or "-"),
            f.get("scanner") or "-",
            location or f.get("cwe") or "-",
            SEV_LABELS.get(sev_key, sev_key.capitalize()),
            f"{round((conf or 0) * 100)}%" if conf is not None else "-",
        ], alt=(i % 2 == 1), left_cols={1, 3, 4, 6})

        # Color severity
        hex_bg, hex_fg = SEV_COLORS.get(sev_key, ("7F8C8D", "FFFFFF"))
        _colored_cell(ws, row, 7, SEV_LABELS.get(sev_key, sev_key.capitalize()), hex_bg, hex_fg)

        ws.row_dimensions[row].height = 22; row += 1

    _data_row(ws, row, [""] * N); ws.row_dimensions[row].height = 18; row += 1

    for j, (_, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


# ─── Tabla 17 — Matriz de escenarios iniciales ─────────────────────────────

def _build_tabla17_scenarios(ws: Any, scenarios: list[dict], asset_map: dict, project_name: str) -> None:
    """
    1 fila = 1 escenario.
    Muestra la agrupacion de hallazgos en escenarios (antes de evaluar impacto).
    """
    ws.title = "T17 Escenarios Iniciales"
    COLS = [
        ("ID Escenario",         14),
        ("ID Activo",            12),
        ("Activo",               26),
        ("Cod. Hallazgos",       18),
        ("N Hallazgos",          12),
        ("Categoria de amenaza", 38),
        ("Proceso de negocio",   26),
    ]
    N = len(COLS)
    row = 1

    _section_title(ws, row, "TFM - WicScan Risk Manager", N)
    ws.row_dimensions[row].height = 22; row += 1

    _section_title(ws, row,
        "Tabla 17. Matriz de escenarios iniciales",
        N, fill=SUBHEAD_FILL)
    ws.row_dimensions[row].height = 20; row += 1

    _note_row(ws, row,
        f"Proyecto: {project_name}   |   {len(scenarios)} escenarios analizados   |   "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", N)
    ws.row_dimensions[row].height = 16; row += 1

    _header_row(ws, row, [c[0] for c in COLS])
    ws.row_dimensions[row].height = 42; row += 1

    for i, s in enumerate(scenarios):
        # Simplify group_key for display (e.g. "uuid_sql_injection" → "sql_injection")
        gk = s.get("group_key") or "-"
        parts = gk.split("_", 1)
        cat_display = parts[1].replace("_", " ").title() if len(parts) > 1 else gk.replace("_", " ").title()

        _data_row(ws, row, [
            s.get("scenario_code") or "-",
            asset_map.get(s.get("asset_name"), "ACT-000"),
            s.get("asset_name") or "Sin activo",
            s.get("finding_codes") or "-",
            s.get("finding_count") or 0,
            cat_display,
            s.get("process_name") or "Sin proceso",
        ], alt=(i % 2 == 1), left_cols={1, 2, 3, 4, 6, 7})

        ws.row_dimensions[row].height = 28; row += 1

    for j, (_, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


# ─── Tabla Escenarios ────────────────────────────────────────────────────────

def _build_escenarios_sheet(ws: Any, scenarios: list[dict], project_name: str) -> None:
    """
    1 fila = 1 escenario (SC-xxx).
    Exportación completa desde /escenarios: P, I, dimensiones, riesgo, tratamientos.
    scenarios dict keys:
        scenario_code, scenario_title, consequence, group_key,
        asset_names (chr10-sep), asset_types, asset_criticalities,
        process_names (chr10-sep), finding_count,
        probability, prob_level, probability_rationale,
        impact, impact_level,
        impact_operational, impact_financial, impact_normative, impact_reputational,
        risk_code, risk_level, risk_status,
        treatment_types, treatment_count, status
    """
    ws.title = "T18 Escenarios de Riesgo"
    COLS = [
        ("ID Escenario",         14),
        ("Titulo escenario",     42),
        ("Consecuencia",         40),
        ("Activos",              26),
        ("Proceso de negocio",   26),
        ("Hallazgos",            20),
        ("P (1-5)",              8),
        ("Nivel P",             14),
        ("Justificacion P",     40),
        ("I (1-5)",              8),
        ("Nivel I",             14),
        ("Justificacion I",     40),
        ("P x I",                7),
    ]
    N = len(COLS)
    row = 1

    _section_title(ws, row, "TFM - WicScan Risk Manager", N)
    ws.row_dimensions[row].height = 22; row += 1

    _section_title(ws, row,
        "Tabla 18. Matriz de escenarios de riesgo",
        N, fill=SUBHEAD_FILL)
    ws.row_dimensions[row].height = 20; row += 1

    _note_row(ws, row,
        f"Proyecto: {project_name}   |   {len(scenarios)} escenarios   |   "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", N)
    ws.row_dimensions[row].height = 16; row += 1

    _header_row(ws, row, [c[0] for c in COLS])
    ws.row_dimensions[row].height = 42; row += 1

    for i, s in enumerate(scenarios):
        p_val  = s.get("probability")
        i_val  = s.get("impact")
        pxi    = (p_val * i_val) if (p_val and i_val) else None
        pl_key = (s.get("prob_level") or "").lower()
        il_key = (s.get("impact_level") or "").lower()

        asset_names  = s.get("asset_names") or "Sin activo"
        process_names = s.get("process_names") or "Sin proceso"

        # Fallback for impact_rationale: show BIA dimension labels when field is empty
        impact_rat = s.get("impact_rationale") or ""
        if not impact_rat:
            dims = []
            for key, label in [
                ("impact_operational", "Operativo"),
                ("impact_financial",   "Financiero"),
                ("impact_normative",   "Normativo"),
                ("impact_reputational","Reputacional"),
            ]:
                val = s.get(key)
                if val:
                    dims.append(f"{label}: {val}")
            impact_rat = " | ".join(dims) if dims else "-"

        # Height scales with the number of assets or processes (1 line = 18pt)
        n_lines = max(
            len(asset_names.split("\n")),
            len(process_names.split("\n")),
        )
        row_h = max(30, n_lines * 18)

        _data_row(ws, row, [
            s.get("scenario_code") or "-",
            s.get("scenario_title") or "-",
            s.get("consequence") or "-",
            asset_names,
            process_names,
            s.get("finding_codes") or str(s.get("finding_count") or 0),
            p_val or "-",
            PROB_LEVEL_LABELS.get(pl_key, pl_key.capitalize()) if pl_key else "-",
            s.get("probability_rationale") or "-",
            i_val or "-",
            IMPACT_LEVEL_LABELS.get(il_key, il_key.capitalize()) if il_key else "-",
            impact_rat,
            pxi if pxi is not None else "-",
        ], alt=(i % 2 == 1), left_cols={1, 2, 3, 4, 5, 6})

        # Color impact level (col 11)
        if il_key and il_key in IMPACT_LEVEL_COLORS:
            ibg, ifg = IMPACT_LEVEL_COLORS[il_key]
            _colored_cell(ws, row, 11, IMPACT_LEVEL_LABELS.get(il_key, il_key.capitalize()), ibg, ifg)

        ws.row_dimensions[row].height = row_h; row += 1

    _data_row(ws, row, [""] * N); ws.row_dimensions[row].height = 18; row += 2

    _section_title(ws, row, "GLOSARIO DE ESTADOS DE ESCENARIO", N, fill=SUBHEAD_FILL)
    ws.row_dimensions[row].height = 18; row += 1
    for term, desc in [
        ("pending",          "Escenario identificado, pendiente de evaluacion de probabilidad"),
        ("prob_assessed",    "Probabilidad asignada por la IA o manualmente"),
        ("impact_assessed",  "Impacto y dimensiones asignados — listo para generar ficha RN-xxx"),
        ("P x I",            "Score ISO 31000: >= 15 Critico, >= 9 Alto, >= 4 Medio, < 4 Bajo"),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        c = ws.cell(row=row, column=1, value=f"  {term}: {desc}")
        c.fill = _fill("F5F5F5"); c.font = Font(name="Calibri", size=9, color="333333")
        c.alignment = _left(); c.border = _border()
        ws.row_dimensions[row].height = 16; row += 1

    for j, (_, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


# ─── Public API ───────────────────────────────────────────────────────────────

def build_bia_excel(processes: list[dict], project_name: str) -> io.BytesIO:
    """Tabla 11 (BIA-1) + Tabla 12 (BIA-2) — descarga desde /contexto/bia."""
    wb = Workbook()
    wb.remove(wb.active)
    _build_bia1(wb.create_sheet(), processes, project_name)
    _build_bia2(wb.create_sheet(), processes, project_name)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_vuln_excel(
    assets: list[dict],
    findings: list[dict],
    scenarios: list[dict],
    project_name: str,
) -> io.BytesIO:
    """Tablas 14, 16, 17 — descarga desde /scans."""
    asset_map = {a.get("name"): f"ACT-{(i+1):03d}" for i, a in enumerate(assets)}

    wb = Workbook()
    wb.remove(wb.active)
    _build_tabla14(wb.create_sheet(), assets, project_name)
    _build_tabla16_findings(wb.create_sheet(), findings, asset_map, project_name)
    _build_tabla17_scenarios(wb.create_sheet(), scenarios, asset_map, project_name)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_scenarios_excel(scenarios: list[dict], project_name: str) -> io.BytesIO:
    """Tabla de Escenarios de Riesgo — descarga desde /escenarios."""
    wb = Workbook()
    wb.remove(wb.active)
    _build_escenarios_sheet(wb.create_sheet(), scenarios, project_name)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Tabla 22 — Registro de riesgos de negocio ──────────────────────────────

def _build_riesgos_sheet(ws: Any, risks: list[dict], project_name: str) -> None:
    ws.title = "T22 Registro Riesgos"
    COLS = [
        ("ID de riesgo",                  14),
        ("Nombre del riesgo",             42),
        ("Hallazgos agrupados",           20),
        ("Activo afectado",               26),
        ("Proceso de negocio afectado",   26),
        ("Descripcion del riesgo",        60),
        ("Probabilidad",                  20),
        ("Impactos",                      25),
        ("Prioridad",                     18),
        ("Nivel de riesgo",               18),
    ]
    N = len(COLS)
    row = 1

    _section_title(ws, row, "TFM - WicScan Risk Manager", N)
    ws.row_dimensions[row].height = 22; row += 1

    _section_title(ws, row,
        "Tabla 22. Registro de riesgos de negocio",
        N, fill=SUBHEAD_FILL)
    ws.row_dimensions[row].height = 20; row += 1

    _note_row(ws, row,
        f"Proyecto: {project_name}   |   {len(risks)} riesgos registrados   |   "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", N)
    ws.row_dimensions[row].height = 16; row += 1

    _header_row(ws, row, [c[0] for c in COLS])
    ws.row_dimensions[row].height = 42; row += 1

    _PRIORITY_LABELS = {
        "immediate":   "Inmediato",
        "short_term":  "Corto plazo",
        "medium_term": "Mediano plazo",
        "long_term":   "Largo plazo",
    }

    for i, r in enumerate(risks):
        pl_key = (r.get("prob_level") or "").lower()
        rl_key = (r.get("risk_level") or "").lower()

        def _clean(val: str | None) -> str:
            v = (val or "").strip()
            return v if v and v != "-" else ""

        proc      = (r.get("process_names") or "").split("\n")[0].strip() or "el proceso afectado"
        asset     = (r.get("asset_names") or "").split("\n")[0].strip() or "el activo de negocio"
        rto       = r.get("bia_rto_hours")
        i24       = r.get("bia_impact_24h")
        op_lv     = (r.get("impact_operational") or "").lower()
        norm_lv   = (r.get("impact_normative") or "").lower()
        rep_lv    = (r.get("impact_reputational") or "").lower()

        # risk_description: use AI text if present, otherwise build from BIA + impact data
        _conseq_map = {
            "muy alto": "interrumpan o comprometan gravemente",
            "alto":     "afecten significativamente",
            "medio":    "degraden parcialmente",
            "bajo":     "afecten de forma limitada",
        }
        _conseq = _conseq_map.get(op_lv, "afecten")
        _fin_clause = (f" Las consecuencias incluyen pérdidas estimadas de {i24:,.0f} EUR" if i24 else " Las consecuencias incluyen pérdidas económicas") + \
                      (", obligaciones de notificación regulatoria" if norm_lv in ("muy alto", "alto") else "") + \
                      (f" y daño reputacional para {proc}" if rep_lv in ("muy alto", "alto") else "") + "."
        _desc_fallback = (
            f"Debido a debilidades en los controles de seguridad de {asset}, existe el riesgo de que "
            f"actores externos {_conseq} la operación de {proc}."
            + (f" Con un RTO de {rto:.0f}h, una interrupción sostenida supera el umbral de tolerancia operativa." if rto else "")
            + _fin_clause
        )
        risk_desc = _clean(r.get("risk_description")) or _desc_fallback

        asset_names = r.get("asset_ids") or "Sin activo"
        process_names = r.get("process_names") or "Sin proceso"

        n_lines = max(
            len(asset_names.split("\n")),
            len(process_names.split("\n")),
        )
        row_h = max(30, n_lines * 18)

        op_val = IMPACT_LEVEL_LABELS.get((r.get("impact_operational") or "").lower(), r.get("impact_operational") or "-")
        fin_val = IMPACT_LEVEL_LABELS.get((r.get("impact_financial") or "").lower(), r.get("impact_financial") or "-")
        norm_val = IMPACT_LEVEL_LABELS.get((r.get("impact_normative") or "").lower(), r.get("impact_normative") or "-")
        rep_val = IMPACT_LEVEL_LABELS.get((r.get("impact_reputational") or "").lower(), r.get("impact_reputational") or "-")
        impactos_str = f"Operativo: {op_val}\nFinanciero: {fin_val}\nNormativo: {norm_val}\nReputacional: {rep_val}"

        _data_row(ws, row, [
            r.get("risk_code") or "-",
            r.get("risk_title") or "-",
            r.get("finding_codes") or "-",
            asset_names,
            process_names,
            risk_desc,
            PROB_LEVEL_LABELS.get(pl_key, pl_key.capitalize()) if pl_key else "-",
            impactos_str,
            _PRIORITY_LABELS.get((r.get("priority") or "").lower(), (r.get("priority") or "-")),
            RISK_LABELS.get(rl_key, rl_key.capitalize()) if rl_key else "-",
        ], alt=(i % 2 == 1), left_cols={1, 2, 3, 4, 5, 6, 8})

        # Color risk level (col 10)
        if rl_key and rl_key in RISK_COLORS:
            rbg, rfg = RISK_COLORS[rl_key]
            _colored_cell(ws, row, 10, RISK_LABELS.get(rl_key, rl_key.capitalize()), rbg, rfg)

        ws.row_dimensions[row].height = row_h; row += 1

    for j, (_, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


def build_risks_excel(risks: list[dict], project_name: str) -> io.BytesIO:
    """Tabla de Registro de Riesgos — descarga desde /analisis."""
    wb = Workbook()
    wb.remove(wb.active)
    _build_riesgos_sheet(wb.create_sheet(), risks, project_name)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
